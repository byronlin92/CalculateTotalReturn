from __future__ import print_function
from lxml import html
import requests
from openpyxl import load_workbook
import sys




def get_stock_price(ticker):
    page = requests.get('https://ca.finance.yahoo.com/quote/' + ticker)
    tree = html.fromstring(page.content)
    price = tree.xpath('//*[@id="quote-summary"]/div[1]/table/tbody/tr[1]/td[2]/span/text()')
    # print (price[0])
    return price[0]

def calculate_total_return(excel_file):
    # variables
    total_dividends = 0
    total_book_value = 0
    total_market_value = 0
    stock_price = 0

    net_amount_column = None
    activity_type_column = None
    symbol_column = None
    quantity_column = None

    wb = load_workbook(filename=excel_file)
    # wb = load_workbook(filename="excel.xlsx")
    if ("Activities" in wb.sheetnames):
        ws = wb["Activities"]
    else:
        return False

    #get columns of total purchase price, activity type
    for col in range (1, ws.max_column):
        if ws.cell(1,col).value == 'Activity Type':
            activity_type_column = col
        if ws.cell(1,col).value == 'Net Amount':
            net_amount_column = col
        if ws.cell(1,col).value == 'Symbol':
            symbol_column = col
        if ws.cell(1,col).value == 'Quantity':
            quantity_column = col


    for row in range (1, ws.max_row):
        #get dividends
        if (ws.cell(row, activity_type_column).value == 'Dividends'):
            total_dividends += float(ws.cell(row,net_amount_column).value)

        #get book value
        if (ws.cell(row, activity_type_column).value == 'Trades'):
            total_book_value += abs(float(ws.cell(row,net_amount_column).value))
            # print abs(float(ws.cell(row,net_amount_column).value))

        #get market value
        if (ws.cell(row, activity_type_column).value == 'Trades'):
            stock_price = float(get_stock_price(ws.cell(row, symbol_column).value))
            quantity = float(ws.cell(row, quantity_column).value)
            total_market_value += (stock_price * quantity)

    # print "market value: " + str(total_market_value)
    # print "book value: " + str(total_book_value)
    # print "dividends: " + str(total_dividends)

    total_return = str(((total_market_value - total_book_value + total_dividends) / total_book_value)*100)
    total_return = total_return[:4] + "%"
    print('total return is: ' + total_return, file=sys.stderr)
    return total_return
    # print "your total return is: " + total_return

def calculate_total_deposit(excel_file):
    # variables
    total_deposits = 0

    activity_type_column = None
    net_amount_column = None

    wb = load_workbook(filename=excel_file)
    # wb = load_workbook(filename="excel.xlsx")
    ws = wb["Activities"]

    #get columns of total purchase price, activity type
    for col in range (1, ws.max_column):
        if ws.cell(1,col).value == 'Activity Type':
            activity_type_column = col
        if ws.cell(1,col).value == 'Net Amount':
            net_amount_column = col

    for row in range (1, ws.max_row):
        #get total deposits
        if (ws.cell(row, activity_type_column).value == 'Deposits'):
            total_deposits += float(ws.cell(row,net_amount_column).value)

    total_deposits = str(total_deposits)
    print('total deposits: ' + total_deposits, file=sys.stderr)
    # print "total deposits: " + str(total_deposits)
    return total_deposits


# calculate_total_return("excel.xlsx")
# calculate_total_deposit("excel.xlsx")




